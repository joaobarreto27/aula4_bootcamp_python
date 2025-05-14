from openpyxl import Workbook, load_workbook
import os

def calcula_bonus(bonus_percentual: float, valor_bonus: float,salario: float):
     resultado: float = valor_bonus + (salario * bonus_percentual)

     return resultado

def valida_nome_usuario(nome: str):

    nome_usuario_formatado = nome.replace(" ", "").lower()
    if len(nome) == 0:
        print("Erro: O nome não pode estar vazio, tente novamente.")
        return False
    elif not nome_usuario_formatado.isalpha():
        print("Erro: O nome deve conter apenas letras, tente novamente.")
        return False
    else:
        return True
    
def valida_salario(salario: float, bonus_percentual: float):
    if salario <= 0:
        print("Erro: O salário inserido é menor ou igual a zero, tente novamente.")
        return False
    elif bonus_percentual < 0:
        print("Erro: O bônus inserido é menor que 0, tente novamente")
        return False
    else:
        return True
    
def valida_status_bonus_salario(
          validacao_bonus_e_salario:bool, 
          bonus_usuario: float, 
          valor_bonus: float, 
          salario_usuario: float, 
          nome_usuario:str):
    if validacao_bonus_e_salario:
            resultado: float = calcula_bonus(bonus_usuario, valor_bonus, salario_usuario)
            print(f"Olá {nome_usuario}, o seu bônus foi de: {resultado:.2f}")
            return True, resultado
    else:
            print("Erro: Tente novamente")
            return False, None

def junta_dados(nome: str, salario: float, bonus_percentual: float, valor_bonus: float, resultado: float):  
    dicionario:dict = {
        'nome': nome,
        'salario': salario,
        'bonus_percentual': bonus_percentual,
        'valor_bonus': valor_bonus,
        'resultado': resultado
    }
    return dicionario

def salva_dados(dados):
    database = "database.xlsx"

    if not os.path.exists(database):
         wb = Workbook()
         ws = wb.active
         ws.title = "data"
         ws.append(list(dados.keys()))
    else:
         wb = load_workbook(database)
         ws = wb["data"]

    ws.append(list(dados.values()))
    wb.save(database)

    return True

def processar_validacoes():
    try:
        # Constante valor de bonus
        valor_bonus: int = 1000
            
        # Entrada e validação nome
        nome_usuario: str = input("Insira seu nome: ")
        if not valida_nome_usuario(nome=nome_usuario):
            return False
        
        # Entrada e validação do salário
        salario_usuario: float = float(input("Insira seu salário: "))
        bonus_usuario: float = float(input("Insira o percentual de bônus recebido (ex: 0.1 para 10%): "))

        validacao_salario_usuario = valida_salario(
              salario= salario_usuario, 
              bonus_percentual= bonus_usuario
        )
        if not validacao_salario_usuario:
              return False
            
        # Execução do cálculo
        status, resultado = valida_status_bonus_salario(
              validacao_bonus_e_salario=validacao_salario_usuario, 
              bonus_usuario=bonus_usuario, 
              valor_bonus=valor_bonus, 
              salario_usuario=salario_usuario, 
              nome_usuario=nome_usuario
        )
        
        # Dados Tabular

        if not status:
            return False

        else:
            dados = junta_dados(
                nome=nome_usuario, 
                salario=salario_usuario, 
                bonus_percentual=bonus_usuario, 
                valor_bonus=valor_bonus, 
                resultado=resultado
            )
            salva_dados(dados=dados)
            return True

    except ValueError:
                print("Erro: Os valores devem conter somente números, tente novamente.")
                return False
    except Exception as e:
            print(f"Erro inesperado: {e}, contate o suporte.")
            return False